import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('path_to_your_file.xlsx')

# Select the sheets
data_sheet = wb['Sheet1']  # or whatever the name of your data sheet is
mapping_sheet = wb['Sheet2']  # or whatever the name of your mapping sheet is

# Add new columns for Location and Region in the data sheet
data_sheet['B1'] = 'Location'
data_sheet['C1'] = 'Region'

# Create a dictionary from the mapping sheet for faster lookup
mapping_dict = {}
for row in range(2, mapping_sheet.max_row + 1):
    host = mapping_sheet.cell(row=row, column=1).value
    location = mapping_sheet.cell(row=row, column=2).value
    region = mapping_sheet.cell(row=row, column=3).value
    mapping_dict[host] = (location, region)

# Loop through each row in the Host column of the data sheet
for row in range(2, data_sheet.max_row + 1):
    host_name = data_sheet.cell(row=row, column=1).value
    if host_name in mapping_dict:
        data_sheet.cell(row=row, column=2).value = mapping_dict[host_name][0]  # Location
        data_sheet.cell(row=row, column=3).value = mapping_dict[host_name][1]  # Region
    else:
        # If not found, you can add logic for closest match here
        pass

# Save the changes
wb.save('path_to_your_file.xlsx')
