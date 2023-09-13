import openpyxl

def get_value_from_mapping(sheet, column, search_value, return_column):
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=column).value == search_value:
            return sheet.cell(row=row, column=return_column).value
    return None

# Load the tool workbook
tool_wb = openpyxl.load_workbook('path_to_tool_workbook.xlsx')
site_mapping_ws = tool_wb['Sheet2']
hostname_mapping_ws = tool_wb['Sheet3']

# Load the downloaded file
data_wb = openpyxl.load_workbook('path_to_downloaded_file.xlsx')
data_ws = data_wb.active

# Insert new columns for Location and Region after the "Host" column
data_ws.insert_cols(6, 2)
data_ws.cell(row=1, column=6).value = 'Location'
data_ws.cell(row=1, column=7).value = 'Region'

# Loop through each row in the downloaded file
for row in range(2, data_ws.max_row + 1):
    site = data_ws.cell(row=row, column=3).value
    hostname = data_ws.cell(row=row, column=5).value

    # Get Location and Region for Site
    location = get_value_from_mapping(site_mapping_ws, 1, site, 2)
    region = get_value_from_mapping(site_mapping_ws, 1, site, 3)

    # If Location is "Rapid7 Insight Agent", get Location and Region for Hostname
    if location == "Rapid7 Insight Agent":
        location = get_value_from_mapping(hostname_mapping_ws, 1, hostname, 2)
        region = get_value_from_mapping(hostname_mapping_ws, 1, hostname, 3)

    # Update Location and Region columns
    data_ws.cell(row=row, column=6).value = location
    data_ws.cell(row=row, column=7).value = region

# Save the changes to the downloaded file
data_wb.save('path_to_downloaded_file.xlsx')
