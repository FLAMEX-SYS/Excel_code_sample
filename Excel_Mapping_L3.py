import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from concurrent.futures import ThreadPoolExecutor

def get_value_from_mapping(df, search_column, search_value, return_column):
    result = df.loc[df[search_column] == search_value, return_column].values
    return result[0] if len(result) > 0 else None

def process_chunk(chunk):
    results = []
    for _, row in chunk.iterrows():
        site = row['Site']
        hostname = row['Host']

        # Get Location and Region for Site
        location = get_value_from_mapping(site_mapping_df, 'Site', site, 'Location')
        region = get_value_from_mapping(site_mapping_df, 'Site', site, 'Region')

        # If Location is "Rapid7 Insight Agent", get Location and Region for Hostname
        if location == "Rapid7 Insight Agent":
            location = get_value_from_mapping(hostname_mapping_df, 'Hostname', hostname, 'Location')
            region = get_value_from_mapping(hostname_mapping_df, 'Hostname', hostname, 'Region')

        results.append([site, hostname, location, region])

    return results

try:
    # Load the tool workbook (.xlsm)
    tool_wb = load_workbook('path_to_tool_workbook.xlsm', data_only=True)  # Use data_only=True to get cell values instead of formulas
    site_mapping_df = pd.read_excel(tool_wb, sheet_name='Sheet2')
    hostname_mapping_df = pd.read_excel(tool_wb, sheet_name='Sheet3')
except KeyError as e:
    print(f"Error: Sheet not found in tool workbook: {e}")
    # Add additional troubleshooting steps here, such as checking the sheet names or file path

# Define chunk size for processing large datasets
chunk_size = 1000

# Initialize the Excel workbook and worksheet
excel_wb = Workbook()
excel_ws = excel_wb.active

# Add headers to the Excel worksheet
headers = ['Site', 'Host', 'Location', 'Region']
excel_ws.append(headers)

# Read and process the downloaded CSV file in chunks using ThreadPoolExecutor
with ThreadPoolExecutor(max_workers=4) as executor:  # Adjust max_workers based on your system's capabilities
    chunks = pd.read_csv('path_to_downloaded_file.csv', chunksize=chunk_size)
    results = executor.map(process_chunk, chunks)

    for chunk_result in results:
        for row in chunk_result:
            excel_ws.append(row)

# Save the Excel workbook
excel_wb.save('output.xlsx')
